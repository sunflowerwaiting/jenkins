from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from git import Repo
import git
import logging
path= 'D:\GIT\software\libva'

def get_tag_hash(path):
    #logging.basicConfig(level=logging.INFO)
    repo = Repo(path)
    #tag_hash = git.Git('git show-ref --tags')
    #print(tag_hash)

    tags = repo.tags
    for tag in tags:
        hash_no = tag.commit
        print(tag)
        print(hash_no)
        commit_log = repo.git.log('--pretty=format:%n%H%n%cd%n%s%n',
                                  date='format:%Y-%m-%d %H:%M')
        log_list = commit_log.split('\n\n')
        real_log_list = []

        wb = Workbook()
        wb.create_sheet('test')
        ws1 = wb['test']
        for log in log_list:
            log = log.split('\n')
            print(log)
            ws1.append(log)
            wb.save('D:\\GIT\\software\\data.xlsx')
        del ws1['Sheet']


'''
            #print(log)
            if log[1] == hash_no:
                log[0] = tag
                real_log_list.append(log)
                #print(real_log_list)
            else :
                pass


        #return real_log_list
    #tagref = tags[0]
    #print(tagref.tag)# tags may have tag objects carrying additional information
   # print(tagref.commit)  # but they always point to commits

    #git show-ref --tags
    #git log --pretty=format:%D%n%H%n%cd%n%s%n
'''
get_tag_hash(path)