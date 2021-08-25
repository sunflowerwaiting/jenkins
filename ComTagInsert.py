import os

from git import Repo, repo
from git import RemoteProgress
import git
import re
from openpyxl import Workbook, load_workbook


############## clone repository #############
def clone_gitlib(ls, path):
    for val in ls:
        val1 = ls[val]
        for sheet in val1:
            git_path = 'git@192.168.16.6:' + val + '/' + sheet + '.git'
            #git.Git(path).clone(git_path)
            #git.Git(path).fetch(git_path)
            print(git_path)
            print('clone  ' + val + '\\' + sheet + '  pass')

def fetch_gitlib(allProjects, path):
    # repo = Repo(self.rorepo.working_tree_dir)
    for val in allProjects:
        val1 = allProjects[val]
        for sheet in val1:
            empty_repo = git.Repo.init(os.path.join(path, sheet))
            print(empty_repo)
            origin = empty_repo.create_remote('origin', repo.remotes.origin.url)
            assert origin.exists()
            assert origin == empty_repo.remotes.origin == empty_repo.remotes['origin']
            origin.fetch()
            git_path = 'git@192.168.16.6:' + val + '/' + sheet + '.git'
            #git.Git(path).fetch(git_path)

            print('pull  ' + val + '\\' + sheet + '  pass')


############### new a excel #####################
def init_excel(project, output_filepath):
    wb = Workbook()

    for sheetName in project:
        wb.create_sheet(sheetName)
        title = ['Function', 'Version Num', 'Commit NUM', 'release date', 'release note (bug fix,new feature)']
        ws1 = wb[sheetName]
        ws1.append(title)

    del wb['Sheet']
    wb.save(output_filepath)


############## get commit message #############
def get_commit(path, repoName):
    repo = Repo(path)

    # fetch
    # origin = repo.remotes['origin']
    # origin.fetch()
    # git log --tags --simplify-by-decoration --pretty=format:%D%n%h%n%cd%n%s%n
    commit_log = repo.git.log('--tags',
                              '--simplify-by-decoration',
                              '--pretty=format:%D%n%h%n%cd%n%s%n',
                              '--reverse',
                              date='format:%Y-%m-%d %H:%M')

    log_list = commit_log.split('\n\n')
    real_log_list = []
    for log in log_list:
        log1 = log.split('\n')
        log1[0] = get_tag(log1[0])
        if log1[0]:
            log1.insert(0, repoName)
            real_log_list.append(log1)
    return real_log_list


regx = re.compile(".*tag:\\s*([\\w\\.-]+)\\b.*")


# ############# use re to  match tag ##############
def get_tag(describe):
    tags = []
    m = regx.search(describe)
    while m:
        tags.append(m.group(1))
        m = regx.search(describe, m.end(1))
    return ", ".join(tags) if len(tags) else None
    # print(describe, m.group(1))


############### insert commit to excel #############
def add_commit_log(commits, wb, val):
    ws1 = wb[val]
    # high of row 1
    ws1.row_dimensions[1].height = 20
    ws1.font = 'Times New Roma'
    ws1.column_dimensions['B'].width = 16
    ws1.column_dimensions['D'].width = 17.89
    ws1.column_dimensions['E'].width = 104
    hashes = []
    for index in range(2,ws1.max_row+1):
        hash_num = 'C'+str(index)
        hashes.append(ws1[hash_num].value)


    for commit in commits:
        if commit[2] in hashes:
            pass
        else:
            ws1.append(commit)
            print(commit)


def GetExcelInfo(exname):
    wb = load_workbook(exname)
    ws1 = wb.sheetnames
