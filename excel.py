
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
work_path = 'D:\\GIT\\software\\'
work_exname = 'SW_version_release.xlsx'
store = work_path + work_exname
def excel_table(path):

    wb = load_workbook(store)
    sheets = wb.sheetnames
    for sheet in sheets:
        ws = wb[sheet]
        max= 'E'+'ws.max_column'
        tab = Table(displayName="Table1",
                    ref="A1:max")  # 名称管理器 如E超出数据范围出错：warn("File may not be readable: column headings must be strings.")

        # 'TableStyleLight11' 1-21 还有此样式  "TableStyleMedium9" 1-28  TableStyleDark1  1-11
        #  showFirstColumn=True,
        #  showLastColumn=True, showRowStripes=True, showColumnStripes=True)
        style = TableStyleInfo(name='TableStyleLight13', showFirstColumn=True,
                               showLastColumn=True, showRowStripes=False, showColumnStripes=True)
        tab.tableStyleInfo = style
        ws.add_table(tab)
        wb.save('D:\\GIT\\software\\SW_version_release.xlsx')

excel_table(store)
